"""
load_history.py
───────────────
SiteInsight · Load Wk01–Wk10 WPR Excel files into Supabase.
Run once to backfill history before live pipeline starts.

Usage:
    python load_history.py --folder ./wpr_files --weeks 1 10
"""

import os
import re
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from supabase import create_client, Client
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Column map: Excel header → DB field ─────────────────────
COL_MAP = {
    "Act\nID": "act_id",
    "Phase": "phase",
    "Activity": "activity",
    "Unit": "unit",
    "Total\nScope": "total_scope",

    # =========================
    # THIS WEEK QTY
    # =========================
    "This Wk\nQty": "this_wk_achieved_qty",  # old
    "This Wk\nAchieved\nQty": "this_wk_achieved_qty",  # new
    "This Wk\nPlanned\nQty": "this_wk_planned_qty",

    # =========================
    # CUM ACTUAL QTY
    # =========================
    "Cum.\nActual Qty": "cum_actual_qty",

    # =========================
    # CUM ACTUAL %
    # =========================
    "Cum.\nActual %": "cum_actual_pct",          # old
    "Cum. Actual %": "cum_actual_pct",          # new (space)
    "Cum. Actual\n%": "cum_actual_pct",         # new (line break)
    "Cum.\nActual\n%": "cum_actual_pct",        # variant

    # =========================
    # CUM PLANNED %
    # =========================
    "★ Cum.\nPlanned %": "cum_planned_pct",     # old
    "★ Cum. Planned %": "cum_planned_pct",     # new (space)
    "★ Cum. Planned\n%": "cum_planned_pct",    # new (line break)
    "Cum.\nPlanned\n%": "cum_planned_pct",     # fallback (no star)

    # =========================
    # OTHER COMMON
    # =========================
    "Variance\n(%)": "variance_pct",
    "Wks\nSlip": "weeks_slip",
    "★ Delay\nReason": "delay_reason",
    "Responsible Person": "responsible_person",
    "★ Critical\nPath": "is_critical_path",
    "Baseline\nVersion": "baseline_version",
    "Remarks / Next Week Plan": "remarks",
}


def parse_header(ws) -> dict:
    """Extract project-level metadata from the first 4 rows."""
    row1 = ws.cell(1, 1).value or ""
    row2 = ws.cell(2, 1).value or ""
    row4_cells = {ws.cell(4, c).value: ws.cell(4, c + 1).value
                  for c in range(1, 15, 2) if ws.cell(4, c).value}

    # Parse row 1: "WEEKLY PROGRESS REPORT | Project | Week 09 | Date | Baseline: B1"
    parts = [p.strip() for p in row1.split("|")]
    project_name = parts[1] if len(parts) > 1 else None
    week_label   = next((p for p in parts if p.startswith("Week")), None)
    week_number  = int(re.search(r"\d+", week_label).group()) if week_label else None

    raw_date = next((p.strip() for p in parts if re.search(r"\d{2}\s\w+\-\d{4}", p)), None)
    report_date = None
    if raw_date:
        try:
            report_date = datetime.strptime(raw_date, "%d %b-%Y").date()
        except ValueError:
            pass

    baseline_version = None
    bl_part = next((p for p in parts if "Baseline:" in p), None)
    if bl_part:
        baseline_version = bl_part.replace("Baseline:", "").strip()

    # Row 2: baseline revision notice
    baseline_revised = "BASELINE REVISED" in row2
    baseline_note    = row2.strip() if baseline_revised else None

    # Row 4: contractor, report week, baseline
    contractor = row4_cells.get("Contractor:")
    total_weeks_raw = row4_cells.get("Report Wk:")
    total_weeks = None
    if total_weeks_raw:
        m = re.search(r"of\s*(\d+)", str(total_weeks_raw))
        if m:
            total_weeks = int(m.group(1))

    return {
        "week_number":                week_number,
        "week_label":                 week_label or f"Wk {week_number:02d}",
        "report_date":                str(report_date) if report_date else None,
        "project_name":               project_name,
        "contractor":                 contractor,
        "total_weeks":                total_weeks,
        "baseline_version":           baseline_version,
        "baseline_revised_this_week": baseline_revised,
        "baseline_revision_note":     baseline_note,
    }


def parse_activities(ws, week_number: int) -> list[dict]:
    """Read activity rows (row 7 onward). Skip phase-header rows."""
    rows = list(ws.iter_rows(min_row=5, values_only=True))
# Skip blank rows after header

    # Row 5 is the column header
    raw_headers = [str(h).strip().replace("\r\n", "\n").replace("\r", "\n") if h else "" for h in rows[0]]
    print(f"DEBUG headers: {raw_headers}")

    records = []
    for raw_row in rows[2:]:
        act_id = raw_row[0]
        if not act_id or str(act_id).startswith("  "):
            continue  # phase header row, skip

        row_dict = dict(zip(raw_headers, raw_row))
        record = {"week_number": week_number}
        for excel_col, db_col in COL_MAP.items():
            val = row_dict.get(excel_col)
            # Skip if val is None AND field already has a value
            if val is None and record.get(db_col) is not None:
                continue
            if db_col == "is_critical_path":
                record[db_col] = str(val).strip().upper() == "Y" if val else False
            elif db_col in ("cum_actual_pct", "cum_planned_pct", "variance_pct"):
                if val is not None:
                    record[db_col] = float(val)
                elif db_col not in record:
                    record[db_col] = None
            elif db_col == "weeks_slip":
                if val is not None:
                    record[db_col] = int(val)
                elif db_col not in record:
                    record[db_col] = 0
            elif db_col in ("total_scope", "this_wk_qty", "cum_actual_qty"):
                if val is not None:
                    record[db_col] = float(val)
                elif db_col not in record:
                    record[db_col] = None
            else:
                if val is not None:
                    record[db_col] = str(val).strip()
                elif db_col not in record:
                    record[db_col] = None
        records.append(record)
    return records


def run_dq_checks(activities: list[dict], week_number: int) -> list[dict]:
    """
    Data-quality pre-flight. Returns list of DQ flag dicts.
    Impossible entries are flagged here; they still load but are marked.
    """
    flags = []
    for act in activities:
        aid = act.get("act_id", "?")

        # R-DQ-1: cumulative % cannot exceed 1.0 (100%)
        for field in ("cum_actual_pct", "cum_planned_pct"):
            val = act.get(field)
            if val is not None and val > 1.0:
                flags.append({
                    "week_number":   week_number,
                    "act_id":        aid,
                    "field_name":    field,
                    "raw_value":     str(val),
                    "rule_triggered": "cum_pct_exceeds_1",
                    "severity":      "error",
                })

        # R-DQ-2: quantities cannot be negative
        for field in ("this_wk_qty", "cum_actual_qty", "total_scope"):
            val = act.get(field)
            if val is not None and val < 0:
                flags.append({
                    "week_number":   week_number,
                    "act_id":        aid,
                    "field_name":    field,
                    "raw_value":     str(val),
                    "rule_triggered": "negative_qty",
                    "severity":      "error",
                })

        # R-DQ-3: variance should equal actual - planned
        actual  = act.get("cum_actual_pct")
        planned = act.get("cum_planned_pct")
        stated  = act.get("variance_pct")
        if all(v is not None for v in (actual, planned, stated)):
            expected = round(actual - planned, 4)
            if abs(expected - round(stated, 4)) > 0.01:
                flags.append({
                    "week_number":   week_number,
                    "act_id":        aid,
                    "field_name":    "variance_pct",
                    "raw_value":     f"stated={stated}, expected={expected}",
                    "rule_triggered": "variance_mismatch",
                    "severity":      "warning",
                })

        # R-DQ-4: cum actual cannot exceed total scope
        scope = act.get("total_scope")
        cum   = act.get("cum_actual_qty")
        if scope and cum and cum > scope:
            flags.append({
                "week_number":   week_number,
                "act_id":        aid,
                "field_name":    "cum_actual_qty",
                "raw_value":     f"cum={cum} > scope={scope}",
                "rule_triggered": "cum_exceeds_scope",
                "severity":      "error",
            })

    return flags


def load_wpr_file(filepath: Path, load_type: str = "history") -> dict:
    """Parse one WPR Excel and upsert into Supabase. Returns summary dict."""
    print(f"\n── Loading {filepath.name} ──")
    wb = load_workbook(str(filepath), read_only=True)

    if "WPR Data" not in wb.sheetnames:
        print(f"  [SKIP] No 'WPR Data' sheet found in {filepath.name}")
        return {"file": filepath.name, "status": "skipped"}

    ws = wb["WPR Data"]
    header = parse_header(ws)
    week_number = header["week_number"]

    if week_number is None:
        print(f"  [ERROR] Could not parse week number from {filepath.name}")
        return {"file": filepath.name, "status": "error"}

    print(f"  Week: {week_number} | Baseline: {header['baseline_version']} | "
          f"Revised: {header['baseline_revised_this_week']}")

    # Upsert WPR header
    header_row = {**header, "source_file": filepath.name, "load_type": load_type}
    supabase.table("wpr_headers").upsert(
        header_row, on_conflict="week_number"
    ).execute()

    # Parse and DQ-check activities
    activities = parse_activities(ws, week_number)
    dq_flags   = run_dq_checks(activities, week_number)

    print(f"  Activities parsed: {len(activities)} | DQ flags: {len(dq_flags)}")

    # Upsert activities
    if activities:
        supabase.table("wpr_activities").upsert(
            activities, on_conflict="week_number,act_id"
        ).execute()

    # Insert DQ flags
    if dq_flags:
        supabase.table("dq_flags").insert(dq_flags).execute()
        for f in dq_flags:
            print(f"  [DQ {f['severity'].upper()}] {f['act_id']} · "
                  f"{f['rule_triggered']} · {f['raw_value']}")

    return {
        "file":        filepath.name,
        "week_number": week_number,
        "activities":  len(activities),
        "dq_flags":    len(dq_flags),
        "status":      "ok",
    }


def main():
    parser = argparse.ArgumentParser(description="Load WPR history into Supabase")
    parser.add_argument("--folder", default="./wpr_files",
                        help="Folder containing WPR Excel files")
    parser.add_argument("--weeks", nargs=2, type=int, default=[1, 10],
                        metavar=("START", "END"),
                        help="Week range to load (default: 1 10)")
    args = parser.parse_args()

    folder     = Path(args.folder)
    week_start = args.weeks[0]
    week_end   = args.weeks[1]

    print(f"\n{'='*55}")
    print(f"  SiteInsight · History Loader")
    print(f"  Folder : {folder}")
    print(f"  Weeks  : {week_start} – {week_end}")
    print(f"{'='*55}")

    # Find files matching WPR_WkNN.xlsx pattern
    files = sorted(folder.glob("WPR_Wk*.xlsx"))
    target_files = []
    for f in files:
        m = re.search(r"Wk(\d+)", f.name)
        if m:
            wk = int(m.group(1))
            if week_start <= wk <= week_end:
                target_files.append(f)

    if not target_files:
        print(f"\n[ERROR] No WPR_WkNN.xlsx files found in {folder} "
              f"for weeks {week_start}–{week_end}")
        return

    print(f"\nFiles to load: {len(target_files)}")
    results = []
    for f in target_files:
        result = load_wpr_file(f, load_type="history")
        results.append(result)

    # Summary
    ok      = sum(1 for r in results if r.get("status") == "ok")
    errors  = sum(1 for r in results if r.get("status") == "error")
    total_dq = sum(r.get("dq_flags", 0) for r in results)

    print(f"\n{'='*55}")
    print(f"  Done. Loaded: {ok} | Errors: {errors} | Total DQ flags: {total_dq}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
